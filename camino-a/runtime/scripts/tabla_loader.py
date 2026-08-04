"""tabla_loader.py — load the CAMINO_N configuration table (xlsx).

The tabla is the source of truth for configuration (OT section 4 D1): changing
a model is editing a cell, never touching code. This module mirrors the
canon_loader.py pattern (dataclass bundle + light validation, no heavy deps)
but reads .xlsx instead of JSON, using openpyxl which is already a dependency
of the runtime (see quality_log delta renderers).

It loads three sheets into typed structures:
  - CAMINO_N_v1_1 : the active route assignments per step/cycle/puesto
  - MODELOS       : the model catalog (route_id -> provider/family/cost/timeout)
  - LOOPS         : the loop policy per step (interno, loop_slot, vuelve_a, topes)

Route/model identities that are not declared in the canon are handled by
contradiccion.py (OT section 8), NOT here: this loader records what the tabla
says and lets the runner classify any disagreement.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional


# Canonical sheet names. The loader refuses to guess: a renamed sheet is a
# configuration error, not something to paper over. The assignments sheet is
# looked up in CANDIDATES order (v1.2 first): the v1.2 tabla replaced the
# assignments sheet and kept the v1.1 context sheets (LEEME, declarado).
SHEET_CAMINO = "CAMINO_N_v1_2"
SHEET_CAMINO_CANDIDATES = ("CAMINO_N_v1_2", "CAMINO_N_v1_1")
SHEET_MODELOS = "MODELOS"
SHEET_LOOPS = "LOOPS"

# tipo_ruta vocabulary (LISTAS sheet). Hardcoding the vocabulary is fine; the
# VALUES still come from the tabla cell.
TIPO_PRIMARIA = "primaria"
TIPO_PARALELA = "paralela"
TIPO_CARRERA = "carrera"
TIPO_FALLBACK = "fallback"
TIPO_CONDICIONAL = "condicional"
TIPOS_VALIDOS = frozenset({
    TIPO_PRIMARIA, TIPO_PARALELA, TIPO_CARRERA, TIPO_FALLBACK, TIPO_CONDICIONAL,
})

NO_CONSTA = "NO_CONSTA"

# Placeholder rows of the CAMINO_N_v1_2 sheet ("Columna1".."ColumnaN").
_PLACEHOLDER_STEP = re.compile(r"^columna\d+$", re.IGNORECASE)


class TablaError(RuntimeError):
    """Raised on unreadable/missing/malformed tabla. -> INPUT_NOT_FOUND_TABLA."""


@dataclass(frozen=True)
class ModelSpec:
    route_id: str
    modelo_exacto: str
    provider_id: str
    provider_name: str
    cost_class: str
    familia: str
    timeout_s: float
    manos: str            # "relee" | "ejecuta"
    modo_agente: str      # "simple" | "nativo_swarm" | "externo_por_eje"
    notas: str = ""

    @property
    def has_hands(self) -> bool:
        """True for rutas con manos: internal loop runs TESTS between steps."""
        return str(self.manos).strip().lower() == "ejecuta"


@dataclass(frozen=True)
class RouteAssignment:
    """One row of CAMINO_N_v1_1: a route placed in a step/puesto/orden."""
    step: int
    ciclo: str
    puesto: str
    capacidades: str
    tipo_ruta: str
    orden: int
    route_id: str
    on_unavailable: str
    provider_name: str
    cost_class: str
    familia: str
    verificacion: str
    latencia_s: float
    timeout_s: float
    manos: str
    fallback_real: str   # "REAL" | "COSMETICO" | "MECANICO" | "" 
    notas: str = ""

    @property
    def is_active(self) -> bool:
        # Inactive rows are marked REVISAR in chk_ruta or have on_unavailable
        # that disables them; the active sheet already excludes disabled rows
        # (LEEME: "Solo puestos ACTIVOS"). Empty route_id = human checkpoint.
        return bool(self.route_id) and self.route_id != NO_CONSTA


@dataclass(frozen=True)
class LoopSpec:
    """One row of LOOPS: loop policy for a step."""
    step: int
    puesto: str
    habilitado: bool
    interno_pasos: int                 # 3 = propone/audita/corrige in one call
    interno_modo: str                  # "una invocacion" | "invocaciones separadas"
    loop_slot: Optional[int]           # slot this step can loop back to
    vuelve_a: Optional[int]            # destination slot
    clase_que_dispara: str             # defect class that triggers the loop
    al_agotarse: str                   # "advance_with_debt" | "restart_big_loop" | ...
    contador_persiste: bool
    tope_ejec: int                     # max executions before exhausting this level


@dataclass(frozen=True)
class TablaConfig:
    tabla_path: Path
    models: dict[str, ModelSpec] = field(default_factory=dict)        # by route_id
    assignments: tuple[RouteAssignment, ...] = ()
    loops: dict[int, LoopSpec] = field(default_factory=dict)         # by step
    # Which assignments sheet was actually read (SHEET_CAMINO_CANDIDATES).
    # Empty for synthetic configs (tests); the generator projects it as-is.
    sheet_camino: str = ""

    def routes_for_step(self, step: int) -> list[RouteAssignment]:
        return [a for a in self.assignments if a.step == step]

    def routes_for_puesto(self, step: int, puesto: str) -> list[RouteAssignment]:
        return [a for a in self.assignments if a.step == step and a.puesto == puesto]

    def model(self, route_id: str) -> Optional[ModelSpec]:
        return self.models.get(route_id)

    def loop_for(self, step: int) -> Optional[LoopSpec]:
        return self.loops.get(step)


# --------------------------------------------------------------------------- #
# Coercion helpers: the tabla is human-edited, so cells are strings, NaN, or
# occasionally numbers. Be lenient on read, strict on meaning.
# --------------------------------------------------------------------------- #
def _s(value: Any, default: str = "") -> str:
    if value is None:
        return default
    s = str(value).strip()
    if s.lower() in {"nan", "none", "null"}:
        return default
    return s


def _i(value: Any, default: int = 0) -> int:
    try:
        if value is None or (isinstance(value, float) and value != value):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _f(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or (isinstance(value, float) and value != value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _bool(value: Any) -> bool:
    return _s(value).upper() in {"SI", "TRUE", "1", "YES", "Y"}


def _open_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise TablaError("openpyxl is required to read the tabla (.xlsx)") from exc
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise TablaError(f"unreadable tabla {path}: {exc}") from exc


def _header_map(ws) -> dict[str, int]:
    """Map header text -> column index from the first non-empty row."""
    mapping: dict[str, int] = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(values_only=True))):
        if cell is None:
            continue
        key = str(cell).strip()
        if key and key not in mapping:
            mapping[key] = col_idx
    return mapping


def _rows(ws, headers: dict[str, int], min_cols: int = 2):
    it = ws.iter_rows(values_only=True)
    next(it, None)  # skip header
    for row in it:
        if row is None:
            continue
        # Skip empty / separator rows.
        non_empty = sum(1 for c in row[:max(min_cols, 1)] if c not in (None, ""))
        if non_empty == 0:
            continue
        yield row


def _row_get(row: tuple, headers: dict[str, int], name: str, default: Any = "") -> Any:
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def _parse_models(ws) -> dict[str, ModelSpec]:
    headers = _header_map(ws)
    out: dict[str, ModelSpec] = {}
    for row in _rows(ws, headers):
        rid = _s(_row_get(row, headers, "route_id"))
        if not rid or rid == NO_CONSTA:
            continue
        spec = ModelSpec(
            route_id=rid,
            modelo_exacto=_s(_row_get(row, headers, "modelo_exacto")),
            provider_id=_s(_row_get(row, headers, "provider_id"), default=rid),
            provider_name=_s(_row_get(row, headers, "provider_name")),
            cost_class=_s(_row_get(row, headers, "cost_class"), NO_CONSTA),
            familia=_s(_row_get(row, headers, "familia"), NO_CONSTA),
            timeout_s=_f(_row_get(row, headers, "timeout_s")),
            manos=_s(_row_get(row, headers, "manos"), "relee"),
            modo_agente=_s(_row_get(row, headers, "modo_agente"), "simple"),
            notas=_s(_row_get(row, headers, "notas")),
        )
        out[rid] = spec
    return out


def _parse_assignments(ws) -> list[RouteAssignment]:
    headers = _header_map(ws)
    out: list[RouteAssignment] = []
    for row in _rows(ws, headers):
        step_raw = _s(_row_get(row, headers, "step"))
        if _PLACEHOLDER_STEP.match(step_raw):
            # The v1.2 sheet keeps a placeholder second-header row
            # ("Columna1".."ColumnaN"); it is sheet scaffolding, not data.
            # Only THIS shape is skipped: any other unknown tipo_ruta still
            # raises below.
            continue
        step_val = _i(step_raw)
        if step_val <= 0:
            # No positive step: either trailing prose (the hoja carries
            # Mariano's annotations below the data, LEEME-declared) or a
            # malformed row. Prose carries NO config content; if any config
            # cell is populated this is a real error — refuse to guess.
            tipo_chk = _s(_row_get(row, headers, "tipo_ruta"))
            route_chk = _s(_row_get(row, headers, "route_id"))
            ciclo_chk = _s(_row_get(row, headers, "ciclo"))
            if tipo_chk or route_chk or ciclo_chk:
                raise TablaError(
                    "fila sin step valido pero con datos de config en la "
                    f"hoja de asignaciones: {list(row[:8])!r}")
            continue
        tipo = _s(_row_get(row, headers, "tipo_ruta")).lower()
        if tipo and tipo not in TIPOS_VALIDOS:
            # An unknown tipo_ruta is a real config error: refuse to guess.
            raise TablaError(
                f"tipo_ruta desconocido '{tipo}' en step "
                f"{_s(_row_get(row, headers, 'step'))}"
            )
        route_id = _s(_row_get(row, headers, "route_id"))
        out.append(RouteAssignment(
            step=_i(_row_get(row, headers, "step")),
            ciclo=_s(_row_get(row, headers, "ciclo"), NO_CONSTA),
            puesto=_s(_row_get(row, headers, "puesto"), NO_CONSTA),
            capacidades=_s(_row_get(row, headers, "capacidades"), NO_CONSTA),
            tipo_ruta=tipo or TIPO_PRIMARIA,
            orden=_i(_row_get(row, headers, "orden")),
            route_id=route_id,
            on_unavailable=_s(_row_get(row, headers, "on_unavailable"), "SKIP_STEP"),
            provider_name=_s(_row_get(row, headers, "provider_name"), NO_CONSTA),
            cost_class=_s(_row_get(row, headers, "cost_class"), NO_CONSTA),
            familia=_s(_row_get(row, headers, "familia"), NO_CONSTA),
            verificacion=_s(_row_get(row, headers, "verificacion"), NO_CONSTA),
            latencia_s=_f(_row_get(row, headers, "latencia_s")),
            timeout_s=_f(_row_get(row, headers, "timeout_s")),
            manos=_s(_row_get(row, headers, "manos"), "relee"),
            fallback_real=_s(_row_get(row, headers, "fallback_real")),
            notas=_s(_row_get(row, headers, "notas")),
        ))
    return out


def _parse_loops(ws) -> dict[int, LoopSpec]:
    headers = _header_map(ws)
    out: dict[int, LoopSpec] = {}
    for row in _rows(ws, headers):
        step = _i(_row_get(row, headers, "step"))
        if not step:
            continue
        loop_slot = _row_get(row, headers, "loop_slot")
        vuelve = _row_get(row, headers, "vuelve_a")
        out[step] = LoopSpec(
            step=step,
            puesto=_s(_row_get(row, headers, "puesto"), NO_CONSTA),
            habilitado=_bool(_row_get(row, headers, "habilitado")),
            interno_pasos=_i(_row_get(row, headers, "interno_pasos"), default=3),
            interno_modo=_s(_row_get(row, headers, "interno_modo"), "una invocacion"),
            loop_slot=_i(loop_slot) if loop_slot not in (None, "") else None,
            vuelve_a=_i(vuelve) if vuelve not in (None, "") else None,
            clase_que_dispara=_s(_row_get(row, headers, "clase_que_dispara")),
            al_agotarse=_s(_row_get(row, headers, "al_agotarse"), "advance_with_debt"),
            contador_persiste=_bool(_row_get(row, headers, "contador_persiste"), ),
            tope_ejec=_i(_row_get(row, headers, "tope_ejec"), default=3),
        )
    return out


def load_tabla(path: Path | str) -> TablaConfig:
    """Load and lightly validate the CAMINO_N tabla.

    Mirrors canon_loader.load_canon: find -> read -> validate -> frozen bundle.
    """
    p = Path(path)
    if not p.is_file():
        raise TablaError(f"tabla no encontrada: {p} (-> INPUT_NOT_FOUND_TABLA)")
    wb = _open_workbook(p)
    sheet_camino = next(
        (s for s in SHEET_CAMINO_CANDIDATES if s in wb.sheetnames), None)
    if sheet_camino is None:
        raise TablaError(
            f"tabla {p.name} sin hoja de asignaciones: probadas "
            f"{list(SHEET_CAMINO_CANDIDATES)}, presentes {wb.sheetnames}")
    missing = [s for s in (SHEET_MODELOS, SHEET_LOOPS) if s not in wb.sheetnames]
    if missing:
        raise TablaError(f"tabla {p.name} sin hojas requeridas: {missing}")

    models = _parse_models(wb[SHEET_MODELOS])
    assignments = tuple(_parse_assignments(wb[sheet_camino]))
    loops = _parse_loops(wb[SHEET_LOOPS])

    # Light validation: every active assignment must reference a known model,
    # OR be a human checkpoint (empty route_id). Unknown route_id with a
    # declared value is left for contradiccion.py to classify — it may be a
    # new route the canon hasn't absorbed yet (CLASE ASIGNACION, tabla wins).
    for a in assignments:
        if a.is_active and a.route_id not in models:
            # Allowed but flagged: contradiccion.py decides if this is a
            # tabla-new-route (ok) or a typo (still ASIGNACION per the rule).
            continue

    return TablaConfig(
        tabla_path=p,
        models=models,
        assignments=assignments,
        loops=loops,
        sheet_camino=sheet_camino,
    )


def find_tabla(intercambio_root: Path, repo_fallback: Path = Path("/tmp/camino-n")) -> Path:
    """Locate the tabla. OT section 0: try the local Intercambio first, then
    the cloned camino-n repo (the tabla was published there on 2-ago-2026).

    Never raises INPUT_NOT_FOUND_TABLA without trying BOTH locations.
    """
    candidates = [
        intercambio_root / "TABLA_CAMINO_N_v1.2_COMPLETA.xlsx",
        intercambio_root / "intercambio" / "TABLA_CAMINO_N_v1.2_COMPLETA.xlsx",
        repo_fallback / "intercambio" / "TABLA_CAMINO_N_v1.2_COMPLETA.xlsx",
        intercambio_root / "TABLA_CAMINO_N_v1.1.xlsx",
        intercambio_root / "intercambio" / "TABLA_CAMINO_N_v1.1.xlsx",
        repo_fallback / "intercambio" / "TABLA_CAMINO_N_v1.1.xlsx",
    ]
    for cand in candidates:
        if cand.is_file():
            return cand
    raise TablaError(
        "tabla ausente en Intercambio y repo fallback -> INPUT_NOT_FOUND_TABLA; "
        f"intentados: {[str(c) for c in candidates]}"
    )
